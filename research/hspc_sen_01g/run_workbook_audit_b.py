#!/usr/bin/env python3
"""HSPC-SEN-01G amendment B: exhaustive two-codepath workbook audit."""
from __future__ import annotations

import hashlib
import io
import json
import os
import re
import tarfile
import unicodedata
import urllib.request
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

import openpyxl

AMENDMENT_SHA256 = "e4c75bad83cc469a94d379554cb023a1d7aee95cbf89f37f9b6910f0edb5c3ba"
PARENT_PREREG_SHA256 = "cd37691108ba4b3d0ded938423675713c46eccceda561f1f4070e2152d99ad29"
AMENDMENT_A_SHA256 = "24865583ee6e3ef9ae1b2229177caae802f239a29ca628d45c046dc8e0a9c294"
PACKAGE_URL = "https://ftp.ncbi.nlm.nih.gov/pub/pmc/deprecated/oa_package/3a/dd/PMC12208344.tar.gz"
EXPECTED_PACKAGE_SHA256 = "71b74bb792dee0a6f2475c21af94411af9bd997fc248b1eef10ab7bc763ae776"
TARGETS = ["mmc2.xlsx", "mmc3.xlsx", "mmc4.xlsx"]
OUT = Path("research/hspc_sen_01g/workbook_audit_b_results")

TERMS = {
    "identity": [
        "donor", "donor id", "subject", "subject id", "participant", "patient",
        "individual", "recipient", "recipient id", "mouse id", "animal id",
        "sample-to-donor", "donor-to-recipient", "crosswalk",
    ],
    "early": [
        "day 1", "day1", "24 h", "24h", "day 4", "day4", "96 h", "96h",
        "senescence", "cdkn1a", "p21", "il1a", "il6", "il8", "p16",
        "sa-beta-gal", "sa-β-gal",
    ],
    "late": [
        "15 week", "15-week", "week 15", "engraftment", "chimerism",
        "multilineage", "clonal diversity", "barcode diversity", "reconstitution",
        "bone marrow", "bm chimerism", "long-term outcome",
    ],
}


def sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def fetch(url: str) -> bytes:
    request = urllib.request.Request(
        url,
        headers={"User-Agent": "HSPC-SEN-01G-workbook-audit/1.0", "Accept": "*/*"},
    )
    with urllib.request.urlopen(request, timeout=240) as response:
        return response.read()


def normalize(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value)).casefold()
    text = "".join(ch if ch.isalnum() else " " for ch in text)
    return " ".join(text.split())


NORMALIZED_TERMS = {
    family: [(term, normalize(term)) for term in terms]
    for family, terms in TERMS.items()
}


def term_hits(records: list[dict[str, str]]) -> dict[str, list[dict[str, str]]]:
    out: dict[str, list[dict[str, str]]] = {family: [] for family in TERMS}
    for record in records:
        normalized_value = record["normalized_value"]
        padded = f" {normalized_value} "
        for family, terms in NORMALIZED_TERMS.items():
            for original_term, normalized_term in terms:
                if f" {normalized_term} " in padded:
                    out[family].append(
                        {
                            "term": original_term,
                            "sheet": record["sheet"],
                            "coordinate": record["coordinate"],
                            "value": record["value"][:500],
                        }
                    )
    return out


def sequence_hash(records: list[dict[str, str]]) -> str:
    payload = "".join(
        f"{record['sheet']}\t{record['coordinate']}\t{record['normalized_value']}\n"
        for record in records
    )
    return sha256(payload.encode("utf-8"))


def path_openpyxl(data: bytes) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=False)
    records: list[dict[str, str]] = []
    sheet_summaries: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        string_count = 0
        nonempty_count = 0
        max_row = 0
        max_column = 0
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                nonempty_count += 1
                max_row = max(max_row, cell.row)
                max_column = max(max_column, cell.column)
                if not isinstance(value, str) or value.startswith("="):
                    continue
                string_count += 1
                records.append(
                    {
                        "sheet": worksheet.title,
                        "coordinate": cell.coordinate,
                        "value": value,
                        "normalized_value": normalize(value),
                    }
                )
        sheet_summaries.append(
            {
                "sheet": worksheet.title,
                "max_row": max_row,
                "max_column": max_column,
                "nonempty_cell_count": nonempty_count,
                "string_cell_count": string_count,
            }
        )
    return {
        "sheet_names": list(workbook.sheetnames),
        "sheet_summaries": sheet_summaries,
        "string_cell_count": len(records),
        "string_sequence_sha256": sequence_hash(records),
        "records": records,
        "hits": term_hits(records),
    }


def all_text(element: ElementTree.Element) -> str:
    return "".join(node.text or "" for node in element.iter() if node.tag.endswith("}t"))


def path_ooxml(data: bytes) -> dict[str, Any]:
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        names = set(archive.namelist())
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in names:
            root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
            shared_strings = [all_text(si) for si in root.findall(".//{*}si")]

        workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        relationships_root = ElementTree.fromstring(
            archive.read("xl/_rels/workbook.xml.rels")
        )
        rels = {
            relationship.attrib["Id"]: relationship.attrib["Target"]
            for relationship in relationships_root.findall(".//{*}Relationship")
        }

        sheets: list[tuple[str, str]] = []
        for sheet in workbook_root.findall(".//{*}sheet"):
            name = sheet.attrib["name"]
            rel_id = next(
                value for key, value in sheet.attrib.items() if key.endswith("}id")
            )
            target = rels[rel_id].lstrip("/")
            if not target.startswith("xl/"):
                target = "xl/" + target
            parts: list[str] = []
            for part in target.split("/"):
                if part == "..":
                    if parts:
                        parts.pop()
                elif part not in {"", "."}:
                    parts.append(part)
            target = "/".join(parts)
            sheets.append((name, target))

        records: list[dict[str, str]] = []
        sheet_summaries: list[dict[str, Any]] = []
        for sheet_name, target in sheets:
            root = ElementTree.fromstring(archive.read(target))
            string_count = 0
            nonempty_count = 0
            max_row = 0
            max_column = 0
            for cell in root.findall(".//{*}c"):
                coordinate = cell.attrib.get("r", "")
                row_match = re.search(r"(\d+)$", coordinate)
                if row_match:
                    max_row = max(max_row, int(row_match.group(1)))
                col_match = re.match(r"([A-Z]+)", coordinate)
                if col_match:
                    column_index = 0
                    for ch in col_match.group(1):
                        column_index = column_index * 26 + (ord(ch) - 64)
                    max_column = max(max_column, column_index)

                formula = cell.find("{*}f")
                cell_type = cell.attrib.get("t")
                value_node = cell.find("{*}v")
                inline_node = cell.find("{*}is")
                raw_value: str | None = None

                if formula is not None:
                    nonempty_count += 1
                    continue
                if cell_type == "s" and value_node is not None:
                    index = int(value_node.text or "0")
                    raw_value = shared_strings[index]
                elif cell_type == "inlineStr" and inline_node is not None:
                    raw_value = all_text(inline_node)
                elif cell_type == "str" and value_node is not None:
                    raw_value = value_node.text or ""
                elif value_node is not None:
                    nonempty_count += 1
                    continue
                elif inline_node is not None:
                    raw_value = all_text(inline_node)

                if raw_value is None:
                    continue
                nonempty_count += 1
                string_count += 1
                records.append(
                    {
                        "sheet": sheet_name,
                        "coordinate": coordinate,
                        "value": raw_value,
                        "normalized_value": normalize(raw_value),
                    }
                )
            sheet_summaries.append(
                {
                    "sheet": sheet_name,
                    "max_row": max_row,
                    "max_column": max_column,
                    "nonempty_cell_count": nonempty_count,
                    "string_cell_count": string_count,
                }
            )

    return {
        "sheet_names": [name for name, _ in sheets],
        "sheet_summaries": sheet_summaries,
        "string_cell_count": len(records),
        "string_sequence_sha256": sequence_hash(records),
        "records": records,
        "hits": term_hits(records),
    }


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    package = fetch(PACKAGE_URL)
    package_hash = sha256(package)
    if package_hash != EXPECTED_PACKAGE_SHA256:
        raise RuntimeError(
            f"OA package hash mismatch: expected {EXPECTED_PACKAGE_SHA256}, got {package_hash}"
        )

    target_data: dict[str, tuple[str, bytes]] = {}
    with tarfile.open(fileobj=io.BytesIO(package), mode="r:*") as archive:
        for member in archive.getmembers():
            if not member.isfile():
                continue
            base = Path(member.name).name
            if base in TARGETS:
                handle = archive.extractfile(member)
                target_data[base] = (member.name, handle.read() if handle else b"")

    results: dict[str, Any] = {}
    reproduction_ok = True
    for target in TARGETS:
        if target not in target_data:
            results[target] = {"state": "FAILED", "reason": "missing from OA package"}
            reproduction_ok = False
            continue
        member_name, raw = target_data[target]
        a = path_openpyxl(raw)
        b = path_ooxml(raw)
        checks = {
            "sheet_names": a["sheet_names"] == b["sheet_names"],
            "string_cell_count": a["string_cell_count"] == b["string_cell_count"],
            "string_sequence_sha256": (
                a["string_sequence_sha256"] == b["string_sequence_sha256"]
            ),
        }
        state = "REPRODUCED" if all(checks.values()) else "FAILED"
        if state != "REPRODUCED":
            reproduction_ok = False
        hits = a["hits"]
        candidate = bool(hits["identity"] and hits["early"] and hits["late"])
        results[target] = {
            "state": state,
            "member_name": member_name,
            "size_bytes": len(raw),
            "sha256": sha256(raw),
            "checks": checks,
            "openpyxl": {
                "sheet_names": a["sheet_names"],
                "sheet_summaries": a["sheet_summaries"],
                "string_cell_count": a["string_cell_count"],
                "string_sequence_sha256": a["string_sequence_sha256"],
                "hits": hits,
            },
            "ooxml": {
                "sheet_names": b["sheet_names"],
                "sheet_summaries": b["sheet_summaries"],
                "string_cell_count": b["string_cell_count"],
                "string_sequence_sha256": b["string_sequence_sha256"],
                "hits": b["hits"],
            },
            "crosswalk_candidate": candidate,
        }

    any_candidate = any(
        item.get("crosswalk_candidate", False) for item in results.values()
    )
    any_identity = any(
        bool(item.get("openpyxl", {}).get("hits", {}).get("identity"))
        for item in results.values()
    )
    if not reproduction_ok:
        decision = "FAILED"
        parent_gate = "BLOCKED_PENDING_SOURCE_AUDIT"
    elif any_candidate:
        decision = "CROSSWALK_PRESENT"
        parent_gate = "MUST_RERUN"
    elif any_identity:
        decision = "INCONCLUSIVE"
        parent_gate = "BLOCKED"
    else:
        decision = "NO_CROSSWALK"
        parent_gate = "BLOCKED"

    result = {
        "analysis_id": "HSPC-SEN-01G-AMENDMENT-B",
        "amendment_sha256": AMENDMENT_SHA256,
        "parent_prereg_sha256": PARENT_PREREG_SHA256,
        "amendment_a_sha256": AMENDMENT_A_SHA256,
        "execution_state": "EXECUTED",
        "reproduction_state": "REPRODUCED" if reproduction_ok else "FAILED",
        "decision": decision,
        "parent_gate_remains": parent_gate,
        "source": {
            "url": PACKAGE_URL,
            "sha256": package_hash,
            "size_bytes": len(package),
        },
        "workbooks": results,
        "safety_verdict": "PRECLINICAL_SOURCE_AUDIT_ONLY",
        "runner": {
            "repository": os.getenv("GITHUB_REPOSITORY"),
            "run_id": os.getenv("GITHUB_RUN_ID"),
            "commit": os.getenv("GITHUB_SHA"),
        },
    }
    (OUT / "workbook_audit_b.json").write_text(
        json.dumps(result, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )

    table_rows = []
    for name, item in results.items():
        hits = item.get("openpyxl", {}).get("hits", {})
        table_rows.append(
            "| {name} | {state} | {count} | {ident} | {early} | {late} | {candidate} |".format(
                name=name,
                state=item.get("state"),
                count=item.get("openpyxl", {}).get("string_cell_count"),
                ident=len(hits.get("identity", [])),
                early=len(hits.get("early", [])),
                late=len(hits.get("late", [])),
                candidate=item.get("crosswalk_candidate"),
            )
        )

    md = f"""# HSPC-SEN-01G technical amendment B result

- Amendment SHA-256: `{AMENDMENT_SHA256}`
- Parent preregistration SHA-256: `{PARENT_PREREG_SHA256}`
- Execution: `EXECUTED`
- Reproduction: `{result['reproduction_state']}`
- Decision: `{decision}`
- Parent gate: `{parent_gate}`
- OA package SHA-256: `{package_hash}`

| Workbook | State | String cells searched | Identity hits | Early hits | Late hits | Crosswalk candidate |
|---|---|---:|---:|---:|---:|---|
{chr(10).join(table_rows)}

Every non-empty string cell was searched. `openpyxl` and raw OOXML extraction had
to agree on sheet names, string-cell count, and the ordered normalized sequence
hash.

The frozen decision is `{decision}`. The HSPC-SEN-01G parent gate remains
`{parent_gate}`. Numeric sample codes, treatment labels, worksheet positions, and
GEO accessions were not treated as donor identities.

No biological, clinical, treatment, dosing, reproductive, germline, or human-safety
claim follows.
"""
    (OUT / "WORKBOOK_AUDIT_B_RESULT.md").write_text(md, encoding="utf-8")
    hashes = []
    for path in sorted(OUT.iterdir()):
        if path.is_file() and path.name != "SHA256SUMS.txt":
            hashes.append(f"{sha256(path.read_bytes())}  {path.name}")
    (OUT / "SHA256SUMS.txt").write_text("\n".join(hashes) + "\n", encoding="utf-8")

    headline = {
        "execution_state": result["execution_state"],
        "reproduction_state": result["reproduction_state"],
        "decision": decision,
        "parent_gate_remains": parent_gate,
        "workbooks": {
            name: {
                "state": item.get("state"),
                "string_cell_count": item.get("openpyxl", {}).get("string_cell_count"),
                "identity_hits": len(item.get("openpyxl", {}).get("hits", {}).get("identity", [])),
                "early_hits": len(item.get("openpyxl", {}).get("hits", {}).get("early", [])),
                "late_hits": len(item.get("openpyxl", {}).get("hits", {}).get("late", [])),
                "crosswalk_candidate": item.get("crosswalk_candidate"),
            }
            for name, item in results.items()
        },
    }
    print("HSPC_WORKBOOK_AUDIT_B_JSON=" + json.dumps(headline, sort_keys=True))
    if not reproduction_ok:
        raise SystemExit(2)


if __name__ == "__main__":
    main()
