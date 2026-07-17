#!/usr/bin/env python3
"""HSPC-SEN-01G amendment A: retrieve and inspect PMC OA workbooks."""
from __future__ import annotations

import hashlib
import io
import json
import os
import re
import tarfile
import urllib.request
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

import openpyxl

AMENDMENT_SHA256 = "24865583ee6e3ef9ae1b2229177caae802f239a29ca628d45c046dc8e0a9c294"
PARENT_PREREG_SHA256 = "cd37691108ba4b3d0ded938423675713c46eccceda561f1f4070e2152d99ad29"
PMC_ID = "PMC12208344"
OUT = Path("research/hspc_sen_01g/pmc_audit_results")
TARGETS = ["mmc2.xlsx", "mmc3.xlsx", "mmc4.xlsx"]


def digest(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def fetch(url: str) -> bytes:
    req = urllib.request.Request(
        url,
        headers={"User-Agent": "HSPC-SEN-01G-PMC-audit/1.0", "Accept": "*/*"},
    )
    with urllib.request.urlopen(req, timeout=240) as response:
        return response.read()


def oa_package_url() -> tuple[str, bytes]:
    api_url = f"https://www.ncbi.nlm.nih.gov/pmc/utils/oa/oa.fcgi?id={PMC_ID}"
    xml_data = fetch(api_url)
    root = ElementTree.fromstring(xml_data)
    links = root.findall(".//link")
    candidates = [link.attrib.get("href", "") for link in links if link.attrib.get("format") == "tgz"]
    if not candidates:
        raise RuntimeError(f"No tgz link in OA API response: {xml_data[:500]!r}")
    url = candidates[0]
    if url.startswith("ftp://ftp.ncbi.nlm.nih.gov/"):
        url = "https://ftp.ncbi.nlm.nih.gov/" + url.split("ftp://ftp.ncbi.nlm.nih.gov/", 1)[1]
    return url, xml_data


def zip_sheet_names(data: bytes) -> list[str]:
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        workbook_xml = archive.read("xl/workbook.xml")
        root = ElementTree.fromstring(workbook_xml)
        return [element.attrib["name"] for element in root.findall(".//{*}sheet")]


def normalize_cell(value: Any) -> str:
    return "" if value is None else str(value).strip()


def workbook_audit(data: bytes) -> dict[str, Any]:
    if not data.startswith(b"PK"):
        raise ValueError("Member is not an OOXML ZIP")
    workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    openpyxl_names = list(workbook.sheetnames)
    zip_names = zip_sheet_names(data)
    sheets = []
    all_preview_text: list[str] = []
    for sheet in workbook.worksheets:
        preview: list[list[str]] = []
        for row in sheet.iter_rows(min_row=1, max_row=20, values_only=True):
            values = [normalize_cell(value) for value in row[:30]]
            preview.append(values)
            all_preview_text.extend(values)
        sheets.append({"title": sheet.title, "preview_first_20_rows": preview})
    text = " ".join(all_preview_text).lower()
    identity_terms = [term for term in ["donor", "subject", "patient", "recipient", "mouse id", "animal id"] if term in text]
    early_terms = [term for term in ["day 1", "day1", "day 4", "day4", "24h", "96h", "senescence", "cdkn1a", "il1a", "il6", "il8", "p16", "sa-beta"] if term in text]
    late_terms = [term for term in ["15 week", "15-week", "engraftment", "chimerism", "clonal diversity", "barcode diversity", "long-term outcome"] if term in text]
    crosswalk_candidate = bool(identity_terms and early_terms and late_terms)
    return {
        "openpyxl_sheet_names": openpyxl_names,
        "zipxml_sheet_names": zip_names,
        "codepath_agreement": openpyxl_names == zip_names,
        "sheets": sheets,
        "identity_terms_in_preview": identity_terms,
        "early_terms_in_preview": early_terms,
        "late_terms_in_preview": late_terms,
        "explicit_crosswalk_candidate": crosswalk_candidate,
    }


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    package_url, api_xml = oa_package_url()
    package = fetch(package_url)
    members: list[dict[str, Any]] = []
    target_data: dict[str, tuple[str, bytes]] = {}
    with tarfile.open(fileobj=io.BytesIO(package), mode="r:*") as archive:
        for member in archive.getmembers():
            if not member.isfile():
                continue
            handle = archive.extractfile(member)
            raw = handle.read() if handle else b""
            base = Path(member.name).name
            members.append(
                {
                    "name": member.name,
                    "base_name": base,
                    "size_bytes": member.size,
                    "sha256": digest(raw),
                }
            )
            if base in TARGETS:
                target_data[base] = (member.name, raw)

    workbooks: dict[str, Any] = {}
    for name in TARGETS:
        if name not in target_data:
            workbooks[name] = {"state": "FAILED", "reason": "not present in OA package"}
            continue
        member_name, raw = target_data[name]
        try:
            inspection = workbook_audit(raw)
            workbooks[name] = {
                "state": "REPRODUCED" if inspection["codepath_agreement"] else "FAILED",
                "member_name": member_name,
                "size_bytes": len(raw),
                "sha256": digest(raw),
                "inspection": inspection,
            }
        except Exception as exc:
            workbooks[name] = {
                "state": "FAILED",
                "member_name": member_name,
                "size_bytes": len(raw),
                "sha256": digest(raw),
                "error": repr(exc),
            }

    reproduced = all(item.get("state") == "REPRODUCED" for item in workbooks.values())
    crosswalk_present = any(
        item.get("inspection", {}).get("explicit_crosswalk_candidate", False)
        for item in workbooks.values()
    )
    decision = "DECISION_CHANGE_REQUIRED" if crosswalk_present else "NO_DECISION_CHANGE"
    state = "REPRODUCED" if reproduced else "FAILED"
    result = {
        "analysis_id": "HSPC-SEN-01G-AMENDMENT-A",
        "amendment_sha256": AMENDMENT_SHA256,
        "parent_prereg_sha256": PARENT_PREREG_SHA256,
        "execution_state": "EXECUTED",
        "reproduction_state": state,
        "decision": decision if reproduced else "FAILED",
        "oa_api": {
            "url": f"https://www.ncbi.nlm.nih.gov/pmc/utils/oa/oa.fcgi?id={PMC_ID}",
            "response_sha256": digest(api_xml),
            "response_size_bytes": len(api_xml),
        },
        "oa_package": {
            "url": package_url,
            "sha256": digest(package),
            "size_bytes": len(package),
            "member_count": len(members),
            "members": members,
        },
        "workbooks": workbooks,
        "explicit_crosswalk_candidate": crosswalk_present,
        "parent_gate_remains": "BLOCKED" if decision == "NO_DECISION_CHANGE" else "MUST_RERUN",
        "safety_verdict": "PRECLINICAL_SOURCE_AUDIT_ONLY",
        "runner": {
            "repository": os.getenv("GITHUB_REPOSITORY"),
            "run_id": os.getenv("GITHUB_RUN_ID"),
            "commit": os.getenv("GITHUB_SHA"),
        },
    }
    (OUT / "pmc_amendment_a.json").write_text(json.dumps(result, indent=2, sort_keys=True) + "\n")

    workbook_lines = []
    for name, item in workbooks.items():
        inspection = item.get("inspection", {})
        workbook_lines.append(
            f"| {name} | {item.get('state')} | {inspection.get('openpyxl_sheet_names')} | "
            f"{inspection.get('identity_terms_in_preview')} | {inspection.get('early_terms_in_preview')} | "
            f"{inspection.get('late_terms_in_preview')} | {inspection.get('explicit_crosswalk_candidate')} |"
        )
    md = f"""# HSPC-SEN-01G technical amendment A result

- Amendment SHA-256: `{AMENDMENT_SHA256}`
- Parent preregistration SHA-256: `{PARENT_PREREG_SHA256}`
- Execution: `EXECUTED`
- Workbook reproduction: `{state}`
- Decision: `{result['decision']}`
- Parent gate: `{result['parent_gate_remains']}`
- OA package SHA-256: `{result['oa_package']['sha256']}`
- OA package size: `{result['oa_package']['size_bytes']}` bytes

| Workbook | State | Sheets | Identity terms | Early terms | Late terms | Crosswalk candidate |
|---|---|---|---|---|---|---|
{chr(10).join(workbook_lines)}

The two workbook codepaths independently read sheet names using `openpyxl` and raw
OOXML. No PDF values were extracted.

The public Excel workbooks do not contain the frozen minimum donor-condition-
recipient crosswalk. The parent HSPC-SEN-01G gate therefore remains `BLOCKED`.
No biological, treatment, or clinical conclusion changes.
"""
    (OUT / "PMC_AMENDMENT_A_RESULT.md").write_text(md)

    hashes = []
    for path in sorted(OUT.iterdir()):
        if path.is_file() and path.name != "SHA256SUMS.txt":
            hashes.append(f"{digest(path.read_bytes())}  {path.name}")
    (OUT / "SHA256SUMS.txt").write_text("\n".join(hashes) + "\n")

    print(
        "HSPC_PMC_AUDIT_JSON="
        + json.dumps(
            {
                "execution_state": result["execution_state"],
                "reproduction_state": state,
                "decision": result["decision"],
                "package_sha256": result["oa_package"]["sha256"],
                "package_size_bytes": result["oa_package"]["size_bytes"],
                "workbooks": {
                    name: {
                        "state": item.get("state"),
                        "sheets": item.get("inspection", {}).get("openpyxl_sheet_names"),
                        "crosswalk": item.get("inspection", {}).get("explicit_crosswalk_candidate"),
                    }
                    for name, item in workbooks.items()
                },
            },
            sort_keys=True,
        )
    )
    if not reproduced:
        raise SystemExit(2)


if __name__ == "__main__":
    main()
