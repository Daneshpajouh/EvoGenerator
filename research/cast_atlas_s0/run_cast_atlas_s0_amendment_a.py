#!/usr/bin/env python3
from __future__ import annotations

import io
import json
import re
from pathlib import Path
from typing import Any

import openpyxl

import run_cast_atlas_s0 as base

TECHNICAL_AMENDMENT_A_SHA256 = "f0c786af45c1bed6bbfd23da2d9ed2a1184570299125e3b24a716e220f548eb9"
DIAGNOSTICS_A: dict[tuple[str, str], dict[str, Any]] = {}
DIAGNOSTICS_B: dict[tuple[str, str], dict[str, Any]] = {}


def explicit_metadata(sample: dict[str, Any]) -> str:
    values: list[str] = []
    for field in ("title", "description", "source", "characteristics"):
        values.extend(str(value).strip() for value in sample.get(field, []) if str(value).strip())
    return " | ".join(values)


def orientation_a(text: str) -> str:
    lower = text.lower()
    if "telomeric" in lower:
        return "TELOMERIC"
    if "centromeric" in lower:
        return "CENTROMERIC"
    return "NO_ORIENTATION"


def orientation_b(text: str) -> str:
    tokens = set(re.sub(r"[^a-z]+", " ", text.lower()).split())
    if "telomeric" in tokens:
        return "TELOMERIC"
    if "centromeric" in tokens:
        return "CENTROMERIC"
    return "NO_ORIENTATION"


def guides_a(text: str) -> str:
    values: set[str] = set()
    lower = text.lower()
    for match in re.finditer(
        r"\b(?:guide(?:\s+rna)?|sgrna|grna|g)\s*[-_:]?\s*(\d+)(?:\s*\+\s*(\d+))?",
        lower,
    ):
        values.add(match.group(1))
        if match.group(2):
            values.add(match.group(2))
    values.update(re.findall(r"#\s*(\d+)", lower))
    return "+".join(sorted(values, key=int)) if values else "UNKNOWN_GUIDE"


def guides_b(text: str) -> str:
    lower = text.lower()
    values: set[int] = set()
    patterns = [
        r"(?:^|[^a-z0-9])g\s*[-_:]?\s*(\d+)(?:\s*\+\s*(\d+))?",
        r"(?:guide\s*(?:rna)?|sgrna|grna)\s*[-_:]?\s*(\d+)(?:\s*\+\s*(\d+))?",
    ]
    for pattern in patterns:
        for left, right in re.findall(pattern, lower):
            values.add(int(left))
            if right:
                values.add(int(right))
    for number in re.findall(r"[a-z0-9]+#\s*(\d+)", lower):
        values.add(int(number))
    return "+".join(str(value) for value in sorted(values)) if values else "UNKNOWN_GUIDE"


def diagnose(
    store: dict[tuple[str, str], dict[str, Any]],
    accession: str,
    sample: dict[str, Any],
    *,
    condition: str | None,
    control: bool,
    incomplete_reason: str | None,
    metadata: str,
) -> None:
    key = (accession, str(sample.get("accession", "")))
    store[key] = {
        "series": accession,
        "sample_accession": str(sample.get("accession", "")),
        "metadata": metadata,
        "metadata_sha256": base.sha256(metadata.encode("utf-8")),
        "control": control,
        "condition": condition,
        "incomplete_reason": incomplete_reason,
    }


def condition_key_a(accession: str, sample: dict[str, Any]) -> str | None:
    metadata = explicit_metadata(sample)
    control = base.is_control(metadata)
    target = base.extract_target(metadata)
    editor = base.extract_editor_a(metadata)
    guide = guides_a(metadata)
    reason = None
    condition = None
    if control:
        reason = "CONTROL"
    elif target == "UNKNOWN_TARGET":
        reason = "UNKNOWN_TARGET"
    elif editor == "UNKNOWN_EDITOR":
        reason = "UNKNOWN_EDITOR"
    else:
        condition = "|".join(
            [
                base.study_group(accession),
                target,
                editor,
                guide,
                orientation_a(metadata),
                base.extract_day(metadata),
                base.extract_template(metadata),
            ]
        )
    diagnose(
        DIAGNOSTICS_A,
        accession,
        sample,
        condition=condition,
        control=control,
        incomplete_reason=reason,
        metadata=metadata,
    )
    return condition


def condition_key_b(accession: str, sample: dict[str, Any]) -> str | None:
    metadata = explicit_metadata(sample)
    control = base.is_control(metadata)
    target = base.extract_target(metadata)
    editor = base.extract_editor_b(metadata)
    guide = guides_b(metadata)
    reason = None
    condition = None
    if control:
        reason = "CONTROL"
    elif target == "UNKNOWN_TARGET":
        reason = "UNKNOWN_TARGET"
    elif editor == "UNKNOWN_EDITOR":
        reason = "UNKNOWN_EDITOR"
    else:
        condition = "|".join(
            [
                base.study_group(accession),
                target,
                editor,
                guide,
                orientation_b(metadata),
                base.extract_day(metadata),
                base.extract_template(metadata),
            ]
        )
    diagnose(
        DIAGNOSTICS_B,
        accession,
        sample,
        condition=condition,
        control=control,
        incomplete_reason=reason,
        metadata=metadata,
    )
    return condition


def openpyxl_tables(name: str, raw: bytes) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=False, data_only=True)
    tables = []
    for sheet in workbook.worksheets:
        rows = [
            list(row)
            for row in sheet.iter_rows(min_row=1, max_row=min(30, sheet.max_row), values_only=True)
        ]
        index, headers = base.choose_header(rows)
        nonempty_headers = [header for header in headers if header]
        tables.append(
            {
                "path": "A_openpyxl_normal_mode",
                "file": name,
                "sheet": sheet.title,
                "header_row_index_zero_based": index,
                "headers": nonempty_headers,
                "field_map": base.classify_headers(nonempty_headers),
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
            }
        )
    workbook.close()
    return tables


def diagnostics_rows(store: dict[tuple[str, str], dict[str, Any]]) -> list[dict[str, Any]]:
    return [store[key] for key in sorted(store)]


def diagnostic_signature(row: dict[str, Any]) -> tuple[Any, ...]:
    return (
        row["series"],
        row["sample_accession"],
        row["metadata_sha256"],
        row["control"],
        row["condition"],
        row["incomplete_reason"],
    )


def write_amended_outputs() -> None:
    result_path = base.OUT / "CAST_ATLAS_S0_RESULT.json"
    result = json.loads(result_path.read_text(encoding="utf-8"))
    rows_a = diagnostics_rows(DIAGNOSTICS_A)
    rows_b = diagnostics_rows(DIAGNOSTICS_B)
    signatures_a = [diagnostic_signature(row) for row in rows_a]
    signatures_b = [diagnostic_signature(row) for row in rows_b]
    sample_diagnostic_agreement = signatures_a == signatures_b

    incomplete_treated = [
        row
        for row in rows_a
        if not row["control"] and row["condition"] is None
    ]
    controls = [row for row in rows_a if row["control"]]

    result["technical_amendment_a_sha256"] = TECHNICAL_AMENDMENT_A_SHA256
    result["execution_state"] = "REPRODUCED_AFTER_TECHNICAL_AMENDMENT_A"
    result["sample_diagnostic_agreement"] = sample_diagnostic_agreement
    result["sample_diagnostics_path_a"] = rows_a
    result["sample_diagnostics_path_b"] = rows_b
    result["excluded_incomplete_treated_samples"] = incomplete_treated
    result["excluded_incomplete_treated_sample_count"] = len(incomplete_treated)
    result["control_sample_count"] = len(controls)

    if not sample_diagnostic_agreement:
        result["decision"] = "FAILED_REPRODUCTION"
        result["rationale"] = "Independent per-sample condition and exclusion diagnostics disagreed."

    amended_json = base.OUT / "CAST_ATLAS_S0_AMENDMENT_A_RESULT.json"
    amended_json.write_text(json.dumps(result, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    md = [
        "# CAST-ATLAS-S0 — amendment-A result",
        "",
        f"- **Public binding SHA-256:** `{base.BINDING_SHA256}`.",
        f"- **Technical amendment A SHA-256:** `{TECHNICAL_AMENDMENT_A_SHA256}`.",
        f"- **Execution:** `{result.get('execution_state')}`.",
        f"- **Decision:** `{result.get('decision')}`.",
        "- **Safety:** data-readiness only; no editor-safety or treatment claim.",
        "",
        "## Independent reproduction",
        "",
        f"- Per-sample condition/exclusion agreement: `{sample_diagnostic_agreement}`.",
        f"- Condition-set agreement: `{result.get('condition_parser_agreement')}`.",
        f"- Treated-condition count: `{result.get('treated_condition_count')}`.",
        f"- Excluded incomplete treated samples: `{len(incomplete_treated)}`.",
        f"- Controls excluded: `{len(controls)}`.",
        "",
        "## Study census",
        "",
        "| Series | Samples | Treated conditions A | Treated conditions B | Agreement |",
        "|---|---:|---:|---:|---|",
    ]
    for accession in base.SERIES:
        item = result.get("soft_results", {}).get(accession, {})
        md.append(
            f"| `{accession}` | {item.get('sample_count_path_a', 'NA')} | "
            f"{item.get('treated_condition_count_path_a', 'NA')} | "
            f"{item.get('treated_condition_count_path_b', 'NA')} | "
            f"`{item.get('condition_parser_agreement')}` |"
        )
    md += [
        "",
        "## Event-table gate",
        "",
        f"- Compatible studies: `{result.get('compatible_studies')}`.",
        f"- Common classified fields: `{result.get('common_fields')}`.",
        f"- Complete common event key: `{result.get('complete_common_event_key')}`.",
        f"- Missed-event denominator measurable: `{result.get('denominator_measurable')}`.",
        "",
        "## Decision",
        "",
        str(result.get("rationale")),
        "",
        "A public collection of positive breakpoint detections is not sufficient for a "
        "leave-study-out atlas certificate without a compatible event key, enough "
        "independent conditions, and a predictor-independent missed-event denominator.",
    ]
    (base.OUT / "CAST_ATLAS_S0_AMENDMENT_A_RESULT.md").write_text(
        "\n".join(md) + "\n", encoding="utf-8"
    )

    manifest = []
    for path in sorted(base.OUT.glob("*")):
        if path.name != "SHA256SUMS.txt":
            manifest.append(f"{base.sha256(path.read_bytes())}  {path.name}")
    (base.OUT / "SHA256SUMS.txt").write_text("\n".join(manifest) + "\n", encoding="utf-8")

    print(
        json.dumps(
            {
                "decision": result.get("decision"),
                "treated_condition_count": result.get("treated_condition_count"),
                "sample_diagnostic_agreement": sample_diagnostic_agreement,
                "compatible_studies": result.get("compatible_studies"),
                "complete_common_event_key": result.get("complete_common_event_key"),
                "denominator_measurable": result.get("denominator_measurable"),
            },
            sort_keys=True,
        )
    )


def main() -> None:
    base.condition_key_a = condition_key_a
    base.condition_key_b = condition_key_b
    base.openpyxl_tables = openpyxl_tables
    base.main()
    write_amended_outputs()


if __name__ == "__main__":
    main()
