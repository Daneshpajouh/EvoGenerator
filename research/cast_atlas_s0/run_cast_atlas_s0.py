#!/usr/bin/env python3
from __future__ import annotations

import csv
import gzip
import hashlib
import io
import json
import re
import tarfile
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any
from urllib.parse import quote
from urllib.request import Request, urlopen
from xml.etree import ElementTree as ET

import openpyxl

BINDING_SHA256 = "b0667e2bc2c82b704a44c2e075d4592caacf3b07f9a99032069f72a53b43afaf"
OUT = Path("research/cast_atlas_s0/results")
OUT.mkdir(parents=True, exist_ok=True)
USER_AGENT = "CAST-ATLAS-S0-schema-gate/1.0"
SERIES = ["GSE164389", "GSE241780", "GSE254922"]
ARCHIVES = {
    "GSE164389": "GSE164389_RAW.tar",
    "GSE241780": "GSE241780_RAW.tar",
}
MAX_BYTES = 300_000_000


def sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def fetch(url: str, timeout: int = 180) -> tuple[bytes, dict[str, Any]]:
    request = Request(url, headers={"User-Agent": USER_AGENT, "Accept": "*/*"})
    with urlopen(request, timeout=timeout) as response:
        data = response.read(MAX_BYTES + 1)
        if len(data) > MAX_BYTES:
            raise RuntimeError(f"download exceeded {MAX_BYTES} bytes")
        return data, {
            "requested_url": url,
            "final_url": response.geturl(),
            "status": getattr(response, "status", None),
            "headers": dict(response.headers.items()),
            "size_bytes": len(data),
            "sha256": sha256(data),
        }


def geo_prefix(accession: str) -> str:
    return accession[:-3] + "nnn"


def soft_url(accession: str) -> str:
    return (
        "https://ftp.ncbi.nlm.nih.gov/geo/series/"
        f"{geo_prefix(accession)}/{accession}/soft/{accession}_family.soft.gz"
    )


def suppl_url(accession: str, filename: str) -> str:
    return (
        "https://ftp.ncbi.nlm.nih.gov/geo/series/"
        f"{geo_prefix(accession)}/{accession}/suppl/{filename}"
    )


def clean(value: str) -> str:
    return value.strip().strip('"').strip()


def parse_soft_a(text: str) -> list[dict[str, Any]]:
    samples: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for raw in text.splitlines():
        line = raw.rstrip("\r\n")
        if line.startswith("^SAMPLE = "):
            if current is not None:
                samples.append(current)
            current = {
                "accession": clean(line.split("=", 1)[1]),
                "title": [],
                "source": [],
                "characteristics": [],
                "description": [],
            }
            continue
        if current is None or not line.startswith("!Sample_") or " = " not in line:
            continue
        key, value = line.split(" = ", 1)
        value = clean(value)
        if key == "!Sample_title":
            current["title"].append(value)
        elif key.startswith("!Sample_source_name"):
            current["source"].append(value)
        elif key.startswith("!Sample_characteristics"):
            current["characteristics"].append(value)
        elif key.startswith("!Sample_description"):
            current["description"].append(value)
    if current is not None:
        samples.append(current)
    return samples


def parse_soft_b(text: str) -> list[dict[str, Any]]:
    chunks = re.split(r"(?m)^\^SAMPLE\s*=\s*", text)[1:]
    samples: list[dict[str, Any]] = []
    for chunk in chunks:
        lines = chunk.splitlines()
        item = {
            "accession": clean(lines[0]) if lines else "",
            "title": [],
            "source": [],
            "characteristics": [],
            "description": [],
        }
        for line in lines[1:]:
            match = re.match(r"^(!Sample_[^=]+?)\s*=\s*(.*)$", line)
            if not match:
                continue
            key = match.group(1).strip()
            value = clean(match.group(2))
            if key == "!Sample_title":
                item["title"].append(value)
            elif key.startswith("!Sample_source_name"):
                item["source"].append(value)
            elif key.startswith("!Sample_characteristics"):
                item["characteristics"].append(value)
            elif key.startswith("!Sample_description"):
                item["description"].append(value)
        samples.append(item)
    return samples


def canonical_sample(sample: dict[str, Any]) -> dict[str, Any]:
    return {
        "accession": sample["accession"],
        "title": sorted(sample["title"]),
        "source": sorted(sample["source"]),
        "characteristics": sorted(sample["characteristics"]),
        "description": sorted(sample["description"]),
    }


TARGETS = ["CCR5", "FANCF", "RAG1", "VEGFA", "COL7A1", "COL17A1", "LAMA3", "WAS", "TRAC"]
CONTROL_TOKENS = (
    "untreated", "non-treated", "non treated", "negative control",
    "mock", "no nuclease", "control sample",
)


def is_control(title: str) -> bool:
    lower = title.lower()
    return any(token in lower for token in CONTROL_TOKENS) or bool(
        re.search(r"(?:^|[-_\s])ut(?:\d+)?(?:$|[-_\s])", lower)
    )


def extract_target(title: str) -> str:
    upper = title.upper()
    hits = [target for target in TARGETS if target in upper]
    return "+".join(sorted(set(hits))) if hits else "UNKNOWN_TARGET"


def extract_editor_a(title: str) -> str:
    text = title.lower()
    if "talens" in text or "talen" in text:
        return "TALEN"
    if "d10acas9n" in text or "double nickase" in text or "cas9n" in text:
        return "CAS9_NICKASE"
    if "hificas9" in text or "high fidelity" in text:
        return "HIFI_CAS9"
    if "cas9" in text:
        return "CAS9"
    return "UNKNOWN_EDITOR"


def extract_editor_b(title: str) -> str:
    tokens = set(re.sub(r"[^a-z0-9]+", " ", title.lower()).split())
    joined = " ".join(tokens)
    if "talen" in joined or "talens" in joined:
        return "TALEN"
    if "d10acas9n" in joined or "cas9n" in joined or {"double", "nickase"} <= tokens:
        return "CAS9_NICKASE"
    if "hificas9" in joined or ({"high", "fidelity"} <= tokens):
        return "HIFI_CAS9"
    if "cas9" in joined:
        return "CAS9"
    return "UNKNOWN_EDITOR"


def extract_guides_a(title: str) -> str:
    lower = title.lower()
    hits = re.findall(r"\bg(?:uide)?[-_\s]?(\d+)(?:\+(\d+))?", lower)
    flattened = []
    for left, right in hits:
        flattened.append(left)
        if right:
            flattened.append(right)
    if not flattened:
        hashes = re.findall(r"#(\d+)", title)
        flattened.extend(hashes)
    return "+".join(sorted(set(flattened), key=lambda x: int(x))) if flattened else "UNKNOWN_GUIDE"


def extract_guides_b(title: str) -> str:
    normalized = re.sub(r"[^a-z0-9#+]+", " ", title.lower())
    numbers: set[str] = set()
    for token in normalized.split():
        match = re.fullmatch(r"g(\d+)(?:\+(\d+))?", token)
        if match:
            numbers.add(match.group(1))
            if match.group(2):
                numbers.add(match.group(2))
        match = re.fullmatch(r"#(\d+)", token)
        if match:
            numbers.add(match.group(1))
    return "+".join(sorted(numbers, key=int)) if numbers else "UNKNOWN_GUIDE"


def extract_day(title: str) -> str:
    match = re.search(r"\bday\s*[-_]?\s*(\d+)\b", title, re.I)
    return f"DAY{match.group(1)}" if match else "NO_DAY"


def extract_template(title: str) -> str:
    return "SSODN" if "ssodn" in title.lower() else "NO_TEMPLATE"


def study_group(accession: str) -> str:
    return "CAST2021" if accession == "GSE164389" else "DUALCAST2024"


def condition_key_a(accession: str, sample: dict[str, Any]) -> str | None:
    title = " | ".join(sample["title"] + sample["description"])
    if is_control(title):
        return None
    return "|".join(
        [
            study_group(accession),
            extract_target(title),
            extract_editor_a(title),
            extract_guides_a(title),
            extract_day(title),
            extract_template(title),
        ]
    )


def condition_key_b(accession: str, sample: dict[str, Any]) -> str | None:
    title = " ; ".join(sample["description"] + sample["title"])
    if is_control(title):
        return None
    return "|".join(
        [
            study_group(accession),
            extract_target(title),
            extract_editor_b(title),
            extract_guides_b(title),
            extract_day(title),
            extract_template(title),
        ]
    )


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value).strip().lower()).strip()


FIELD_TOKENS: dict[str, list[tuple[str, ...]]] = {
    "condition_identity": [
        ("sample",), ("condition",), ("treatment",), ("target",), ("guide",), ("sgrna",),
    ],
    "reference_assembly": [("assembly",), ("genome", "build"), ("reference",)],
    "bait_identity": [("bait",), ("anchor",), ("on target",), ("ontarget",)],
    "partner_chromosome": [
        ("partner", "chr"), ("translocation", "chr"), ("chromosome",), ("chr",),
    ],
    "partner_coordinate": [
        ("breakpoint",), ("break point",), ("coordinate",), ("position",),
        ("start",), ("end",),
    ],
    "event_class": [
        ("event", "type"), ("event", "class"), ("rearrangement",),
        ("translocation", "type"), ("category",), ("class",),
    ],
    "quantitative_support": [
        ("read",), ("count",), ("support",), ("cpm",), ("frequency",),
        ("percentage",), ("fraction",), ("coverage",),
    ],
    "control_or_validation": [
        ("control",), ("treated",), ("untreated",), ("validation",), ("validated",),
    ],
    "replicate_identity": [("replicate",), ("duplicate",), ("library",), ("sample", "id")],
}


def classify_headers(headers: list[str]) -> dict[str, list[str]]:
    out: dict[str, list[str]] = {}
    for field, groups in FIELD_TOKENS.items():
        matches = []
        for header in headers:
            norm = normalize_header(header)
            if any(all(token in norm for token in group) for group in groups):
                matches.append(header)
        out[field] = matches
    return out


def header_score(row: list[Any]) -> int:
    cells = [str(value).strip() for value in row if value not in (None, "")]
    if not cells:
        return -1
    norms = [normalize_header(value) for value in cells]
    keyword_hits = sum(
        1
        for norm in norms
        if any(
            token in norm
            for token in (
                "sample", "condition", "chr", "chromosome", "position",
                "breakpoint", "read", "count", "event", "type", "target",
                "guide", "translocation", "deletion", "inversion",
            )
        )
    )
    return len(cells) + 3 * keyword_hits


def choose_header(rows: list[list[Any]]) -> tuple[int | None, list[str]]:
    best_index = None
    best_score = -1
    best_headers: list[str] = []
    for index, row in enumerate(rows[:30]):
        score = header_score(row)
        if score > best_score:
            best_score = score
            best_index = index
            best_headers = [str(value).strip() if value is not None else "" for value in row]
    return best_index, best_headers


def openpyxl_tables(name: str, raw: bytes) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    tables = []
    for sheet in workbook.worksheets:
        rows = [
            list(row)
            for row in sheet.iter_rows(min_row=1, max_row=30, values_only=True)
        ]
        index, headers = choose_header(rows)
        nonempty_headers = [h for h in headers if h]
        tables.append(
            {
                "path": "A_openpyxl",
                "file": name,
                "sheet": sheet.title,
                "header_row_index_zero_based": index,
                "headers": nonempty_headers,
                "field_map": classify_headers(nonempty_headers),
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
            }
        )
    return tables


NS = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
      "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
REL_NS = {"p": "http://schemas.openxmlformats.org/package/2006/relationships"}


def column_index(cell_reference: str) -> int:
    letters = re.match(r"[A-Z]+", cell_reference)
    value = 0
    for char in letters.group(0) if letters else "A":
        value = value * 26 + ord(char) - 64
    return value - 1


def raw_ooxml_tables(name: str, raw: bytes) -> list[dict[str, Any]]:
    tables = []
    with zipfile.ZipFile(io.BytesIO(raw)) as archive:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for si in root.findall("a:si", NS):
                shared.append("".join(node.text or "" for node in si.findall(".//a:t", NS)))
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        targets = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels.findall("p:Relationship", REL_NS)
        }
        for sheet in workbook.findall("a:sheets/a:sheet", NS):
            title = sheet.attrib.get("name", "")
            rel_id = sheet.attrib.get(f"{{{NS['r']}}}id")
            target = targets.get(rel_id, "")
            sheet_path = target if target.startswith("xl/") else f"xl/{target.lstrip('/')}"
            if sheet_path not in archive.namelist():
                continue
            root = ET.fromstring(archive.read(sheet_path))
            rows: list[list[Any]] = []
            for row in root.findall(".//a:sheetData/a:row", NS)[:30]:
                values: dict[int, Any] = {}
                for cell in row.findall("a:c", NS):
                    ref = cell.attrib.get("r", "A1")
                    idx = column_index(ref)
                    kind = cell.attrib.get("t")
                    value_node = cell.find("a:v", NS)
                    inline = cell.find("a:is", NS)
                    value: Any = ""
                    if kind == "s" and value_node is not None:
                        pointer = int(value_node.text or "0")
                        value = shared[pointer] if pointer < len(shared) else ""
                    elif kind == "inlineStr" and inline is not None:
                        value = "".join(node.text or "" for node in inline.findall(".//a:t", NS))
                    elif value_node is not None:
                        value = value_node.text or ""
                    values[idx] = value
                width = max(values, default=-1) + 1
                rows.append([values.get(index, "") for index in range(width)])
            index, headers = choose_header(rows)
            nonempty_headers = [h for h in headers if h]
            dimension = root.find("a:dimension", NS)
            tables.append(
                {
                    "path": "B_raw_ooxml",
                    "file": name,
                    "sheet": title,
                    "header_row_index_zero_based": index,
                    "headers": nonempty_headers,
                    "field_map": classify_headers(nonempty_headers),
                    "dimension": dimension.attrib.get("ref") if dimension is not None else None,
                }
            )
    return tables


def text_table(name: str, raw: bytes) -> list[dict[str, Any]]:
    text = raw.decode("utf-8-sig", "replace")
    sample = "\n".join(text.splitlines()[:30])
    delimiter = "\t" if sample.count("\t") >= sample.count(",") else ","
    reader = csv.reader(io.StringIO(sample), delimiter=delimiter)
    rows = [row for row in reader]
    index, headers = choose_header(rows)
    nonempty_headers = [h for h in headers if h]
    return [
        {
            "path": "TEXT",
            "file": name,
            "sheet": None,
            "header_row_index_zero_based": index,
            "headers": nonempty_headers,
            "field_map": classify_headers(nonempty_headers),
            "delimiter": delimiter,
            "line_count": len(text.splitlines()),
        }
    ]


def inspect_archive(accession: str, data: bytes) -> dict[str, Any]:
    members: list[dict[str, Any]] = []
    tables_a: list[dict[str, Any]] = []
    tables_b: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    with tarfile.open(fileobj=io.BytesIO(data), mode="r:*") as archive:
        for member in sorted((m for m in archive.getmembers() if m.isfile()), key=lambda m: m.name):
            handle = archive.extractfile(member)
            raw = handle.read() if handle else b""
            members.append(
                {
                    "name": member.name,
                    "size_bytes": member.size,
                    "sha256": sha256(raw),
                }
            )
            lower = member.name.lower()
            try:
                if lower.endswith(".xlsx"):
                    tables_a.extend(openpyxl_tables(member.name, raw))
                    tables_b.extend(raw_ooxml_tables(member.name, raw))
                elif lower.endswith((".txt", ".tsv", ".csv")):
                    parsed = text_table(member.name, raw)
                    tables_a.extend([{**row, "path": "A_text_csv"} for row in parsed])
                    tables_b.extend([{**row, "path": "B_text_csv"} for row in parsed])
            except Exception as exc:
                errors.append({"member": member.name, "error": repr(exc)})
    return {
        "accession": accession,
        "members": members,
        "tables_path_a": tables_a,
        "tables_path_b": tables_b,
        "errors": errors,
    }


def schema_signature(table: dict[str, Any]) -> tuple[str, str, tuple[str, ...], int | None]:
    return (
        table.get("file") or "",
        table.get("sheet") or "",
        tuple(sorted(normalize_header(header) for header in table.get("headers", []))),
        table.get("header_row_index_zero_based"),
    )


def event_table_candidate(table: dict[str, Any]) -> bool:
    field_map = table.get("field_map", {})
    required = [
        "condition_identity",
        "partner_chromosome",
        "partner_coordinate",
        "quantitative_support",
    ]
    return all(bool(field_map.get(field)) for field in required)


def ena_metadata() -> dict[str, Any]:
    fields = ",".join(
        [
            "study_accession", "sample_accession", "experiment_accession",
            "run_accession", "scientific_name", "instrument_platform",
            "instrument_model", "library_strategy", "library_source",
            "library_selection", "fastq_ftp", "submitted_ftp",
        ]
    )
    url = (
        "https://www.ebi.ac.uk/ena/portal/api/filereport?"
        f"accession=PRJEB106504&result=read_run&fields={quote(fields)}"
        "&format=tsv&download=true"
    )
    try:
        data, meta = fetch(url)
        text = data.decode("utf-8-sig", "replace")
        rows = list(csv.DictReader(io.StringIO(text), delimiter="\t"))
        return {
            "state": "DATA_ACQUIRED",
            "source": meta,
            "row_count": len(rows),
            "columns": list(rows[0]) if rows else [],
            "study_accessions": sorted({row.get("study_accession", "") for row in rows if row.get("study_accession")}),
            "sample_count": len({row.get("sample_accession", "") for row in rows if row.get("sample_accession")}),
            "run_count": len({row.get("run_accession", "") for row in rows if row.get("run_accession")}),
            "processed_event_table_present": False,
        }
    except Exception as exc:
        return {"state": "FAILED", "url": url, "error": repr(exc)}


def main() -> None:
    binding = Path("research/cast_atlas_s0/CAST_ATLAS_S0_PUBLIC_BINDING.md").read_bytes()
    if sha256(binding) != BINDING_SHA256:
        raise SystemExit("public binding SHA-256 mismatch")

    sources: dict[str, Any] = {}
    soft_results: dict[str, Any] = {}
    all_conditions_a: set[str] = set()
    all_conditions_b: set[str] = set()
    acquisitions_failed = False

    for accession in SERIES:
        url = soft_url(accession)
        try:
            data, meta = fetch(url)
            sources[f"{accession}_soft"] = meta
            text = gzip.decompress(data).decode("utf-8-sig", "replace")
            a = parse_soft_a(text)
            b = parse_soft_b(text)
            canonical_a = sorted((canonical_sample(item) for item in a), key=lambda x: x["accession"])
            canonical_b = sorted((canonical_sample(item) for item in b), key=lambda x: x["accession"])
            conditions_a = {key for item in a if (key := condition_key_a(accession, item))}
            conditions_b = {key for item in b if (key := condition_key_b(accession, item))}
            all_conditions_a.update(conditions_a)
            all_conditions_b.update(conditions_b)
            soft_results[accession] = {
                "sample_count_path_a": len(a),
                "sample_count_path_b": len(b),
                "sample_parser_agreement": canonical_a == canonical_b,
                "treated_condition_count_path_a": len(conditions_a),
                "treated_condition_count_path_b": len(conditions_b),
                "condition_parser_agreement": conditions_a == conditions_b,
                "conditions_path_a": sorted(conditions_a),
                "conditions_path_b": sorted(conditions_b),
            }
        except Exception as exc:
            acquisitions_failed = True
            soft_results[accession] = {"state": "FAILED", "url": url, "error": repr(exc)}

    archive_results: dict[str, Any] = {}
    for accession, filename in ARCHIVES.items():
        url = suppl_url(accession, filename)
        try:
            data, meta = fetch(url)
            sources[f"{accession}_archive"] = meta
            archive_results[accession] = inspect_archive(accession, data)
        except Exception as exc:
            acquisitions_failed = True
            archive_results[accession] = {"state": "FAILED", "url": url, "error": repr(exc)}

    ena = ena_metadata()

    parser_agreement = all(
        item.get("sample_parser_agreement") and item.get("condition_parser_agreement")
        for item in soft_results.values()
        if item.get("state") != "FAILED"
    )
    archive_agreement = True
    for result in archive_results.values():
        if result.get("state") == "FAILED":
            continue
        signatures_a = sorted(schema_signature(table) for table in result["tables_path_a"])
        signatures_b = sorted(schema_signature(table) for table in result["tables_path_b"])
        result["schema_parser_agreement"] = signatures_a == signatures_b
        result["schema_signatures_path_a"] = signatures_a
        result["schema_signatures_path_b"] = signatures_b
        archive_agreement = archive_agreement and signatures_a == signatures_b

    conditions_agree = all_conditions_a == all_conditions_b
    treated_condition_count = len(all_conditions_a) if conditions_agree else None

    candidate_tables_by_study: dict[str, list[dict[str, Any]]] = {}
    common_fields_by_study: dict[str, set[str]] = {}
    for accession, result in archive_results.items():
        if result.get("state") == "FAILED":
            continue
        candidates = [
            table for table in result["tables_path_a"] if event_table_candidate(table)
        ]
        candidate_tables_by_study[accession] = candidates
        fields = set()
        for table in candidates:
            for field, matches in table.get("field_map", {}).items():
                if matches:
                    fields.add(field)
        common_fields_by_study[accession] = fields

    compatible_studies = [study for study, tables in candidate_tables_by_study.items() if tables]
    common_fields = (
        set.intersection(*(common_fields_by_study[study] for study in compatible_studies))
        if compatible_studies
        else set()
    )
    full_required_fields = set(FIELD_TOKENS)
    full_common_key = full_required_fields <= common_fields

    denominator_measurable = False
    denominator_evidence: list[dict[str, Any]] = []
    for study, tables in candidate_tables_by_study.items():
        for table in tables:
            fields = table.get("field_map", {})
            if fields.get("control_or_validation") and fields.get("replicate_identity"):
                denominator_measurable = True
                denominator_evidence.append(
                    {"study": study, "file": table.get("file"), "sheet": table.get("sheet")}
                )

    if not parser_agreement or not archive_agreement or not conditions_agree:
        decision = "FAILED_REPRODUCTION"
        rationale = "Independent metadata, condition, or table-schema parsers disagreed."
    elif acquisitions_failed:
        decision = "FAILED_ACQUISITION"
        rationale = "At least one required GEO source failed acquisition."
    elif len(compatible_studies) < 2:
        any_tables = any(
            result.get("tables_path_a")
            for result in archive_results.values()
            if result.get("state") != "FAILED"
        )
        decision = "SCHEMA_BARRIER" if any_tables else "PROCESSED_TABLE_ABSENT"
        rationale = "Fewer than two independent studies exposed a compatible event-level table."
    elif not full_common_key:
        decision = "SCHEMA_BARRIER"
        rationale = "Candidate event tables did not share the complete frozen event key."
    elif treated_condition_count is None or treated_condition_count < 59:
        decision = "SAMPLE_SIZE_BARRIER"
        rationale = (
            f"Only {treated_condition_count} independent treated conditions remained "
            "after conservative collapse; 59 were required."
        )
    elif not denominator_measurable:
        decision = "DENOMINATOR_BARRIER"
        rationale = "No predictor-independent missed-event denominator was measurable."
    else:
        decision = "EXECUTABLE_ATLAS_ROUTE"
        rationale = "Every frozen schema, sample-size, denominator, and reproduction gate passed."

    result = {
        "analysis_id": "CAST-ATLAS-S0",
        "public_binding_sha256": BINDING_SHA256,
        "private_prereg_git_blob_sha": "a7d7d7dac61130abb588b7c2006e1fca4c22d676",
        "execution_state": "EXECUTED",
        "decision": decision,
        "rationale": rationale,
        "sources": sources,
        "soft_results": soft_results,
        "treated_conditions_path_a": sorted(all_conditions_a),
        "treated_conditions_path_b": sorted(all_conditions_b),
        "treated_condition_count": treated_condition_count,
        "condition_parser_agreement": conditions_agree,
        "archive_results": archive_results,
        "candidate_tables_by_study": candidate_tables_by_study,
        "compatible_studies": compatible_studies,
        "common_fields": sorted(common_fields),
        "complete_common_event_key": full_common_key,
        "denominator_measurable": denominator_measurable,
        "denominator_evidence": denominator_evidence,
        "ena_prjeb106504": ena,
        "safety_verdict": "DATA_READINESS_ONLY_NO_EDITOR_OR_TREATMENT_CLAIM",
    }
    result_path = OUT / "CAST_ATLAS_S0_RESULT.json"
    result_path.write_text(json.dumps(result, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    md = [
        "# CAST-ATLAS-S0 — public structural-event schema and power gate",
        "",
        f"- **Public binding SHA-256:** `{BINDING_SHA256}`.",
        "- **Execution:** `EXECUTED`.",
        f"- **Decision:** `{decision}`.",
        "- **Safety:** data-readiness only; no editor-safety or treatment claim.",
        "",
        "## Source and condition census",
        "",
        "| Series | Samples | Parser agreement | Treated conditions A | Treated conditions B |",
        "|---|---:|---|---:|---:|",
    ]
    for accession in SERIES:
        item = soft_results.get(accession, {})
        md.append(
            f"| `{accession}` | {item.get('sample_count_path_a', 'NA')} | "
            f"`{item.get('sample_parser_agreement')}` | "
            f"{item.get('treated_condition_count_path_a', 'NA')} | "
            f"{item.get('treated_condition_count_path_b', 'NA')} |"
        )
    md += [
        "",
        f"- Deduplicated treated conditions across source groups: `{treated_condition_count}`.",
        f"- Condition-parser agreement: `{conditions_agree}`.",
        "",
        "## Archive and table census",
        "",
        "| Study | Archive members | Parsed tables/sheets | Schema parser agreement | Event-table candidates |",
        "|---|---:|---:|---|---:|",
    ]
    for accession, item in archive_results.items():
        if item.get("state") == "FAILED":
            md.append(f"| `{accession}` | NA | NA | `False` | NA |")
            continue
        md.append(
            f"| `{accession}` | {len(item.get('members', []))} | "
            f"{len(item.get('tables_path_a', []))} | "
            f"`{item.get('schema_parser_agreement')}` | "
            f"{len(candidate_tables_by_study.get(accession, []))} |"
        )
    md += [
        "",
        f"- Compatible studies with event-table candidates: `{compatible_studies}`.",
        f"- Common classified fields: `{sorted(common_fields)}`.",
        f"- Complete frozen common event key: `{full_common_key}`.",
        f"- Predictor-independent missed-event denominator measurable: `{denominator_measurable}`.",
        "",
        "## ENA route",
        "",
        f"- PRJEB106504 state: `{ena.get('state')}`.",
        f"- Read-run rows: `{ena.get('row_count')}`.",
        f"- Processed event table present: `{ena.get('processed_event_table_present')}`.",
        "",
        "## Decision",
        "",
        rationale,
        "",
        "A public set of positive CAST-Seq detections is not yet an atlas certificate. "
        "The next build is permitted only if the complete event key, independent "
        "condition floor, and missed-event denominator all survive this gate.",
    ]
    (OUT / "CAST_ATLAS_S0_RESULT.md").write_text("\n".join(md) + "\n", encoding="utf-8")

    lines = []
    for path in sorted(OUT.glob("*")):
        if path.name != "SHA256SUMS.txt":
            lines.append(f"{sha256(path.read_bytes())}  {path.name}")
    (OUT / "SHA256SUMS.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(
        json.dumps(
            {
                "decision": decision,
                "treated_condition_count": treated_condition_count,
                "compatible_studies": compatible_studies,
                "complete_common_event_key": full_common_key,
                "denominator_measurable": denominator_measurable,
            },
            sort_keys=True,
        )
    )


if __name__ == "__main__":
    main()
